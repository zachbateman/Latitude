#!/usr/bin/env python3
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import pandas as pd
import xlsxwriter
import openpyxl
import datetime


# Server File location
loc2 = ('S:\OPERATIONS & ENGINEERING\Well DR Pressure\Dashboard - Wells.xlsm')
# Local file location for off network
# loc2 =('C:\\Users\camoruso\Desktop\Dashboard - Wells.xlsm')

sheetname = "WS Alt Data"

# Login Info from local file (username and password will be module-level variables)
with open("C:\\Users\camoruso\Desktop\password.txt", "r") as login_file:
    contents = login_file.readlines()
    username = contents[0]
    password = contents[1]



def get_workbook_parameters(filename, sheet_name):
    '''
    Open Worksheet with Pandas
    Return row count, Begdate, and Enddate
    '''
    wellsdf = pd.read_excel(filename, sheet_name=sheetname)
    Excelrowcount = len(wellsdf)
    columns = wellsdf.columns
    Datelist = wellsdf[columns[0]].tolist()
    lastdate = Datelist[-1]

    Yesterday = datetime.date.today() + datetime.timedelta(days=-1)
    if lastdate.date() < Yesterday:
        Begdate = lastdate.date() + datetime.timedelta(days=1)
        Begdate = Begdate.strftime('%m/%d/%Y')
        print(Begdate)
        Enddate = Yesterday
        Enddate = Enddate.strftime('%m/%d/%Y')
        print(Enddate)
    else:
        Begdate = "05/02/2019"
        Enddate = "05/07/2019"

    return Excelrowcount, Begdate, Enddate


def df_from_web(Begdate, Enddate):
    '''
    Read data from gasstorage.net
    Return OBAdf DataFrame.
    '''
    driver = webdriver.Chrome('C:\\Users\camoruso\Documents\Helpful Docs\Programming\chromedriver.exe')
    driver.get('https://www.gasstorage.net/WORSHAMSTEED/Operator/index.cfm')

    element = driver.find_element_by_xpath('//*[@id="username"]')
    element.send_keys(username)

    element = driver.find_element_by_xpath('//*[@id="password"]')
    element.send_keys(password)

    element = driver.find_element_by_xpath('//*[@id="frmLogin"]/table[1]/tbody/tr[3]/td/input')
    element.send_keys(Keys.RETURN)

    element = driver.find_element_by_xpath('//*[@id="navmenu"]/table/tbody/tr[14]/td/a')
    element.send_keys(Keys.RETURN)

    element = driver.find_element_by_xpath('//*[@id="NetFacilStart_Disp"]')
    element.clear()
    element.send_keys(Begdate)

    element = driver.find_element_by_xpath('//*[@id="NetFacilEnd_Disp"]')
    element.clear()
    element.send_keys(Enddate)

    element = driver.find_element_by_xpath('//*[@id="content"]/div/div[3]/fieldset[3]/form/table/tbody/tr/td[5]/input')
    element.send_keys(Keys.RETURN)

    table = driver.find_element_by_xpath('//*[@id="content"]/div/div[3]/table/tbody')
    rows = table.find_elements_by_tag_name("tr")

    x = []
    for row in rows:
        x.append([td.text for td in row.find_elements_by_tag_name("td")])
    OBAdf = pd.DataFrame(x)

    # Close webdriver
    driver.close()
    driver.quit()

    return OBAdf


def update_workbook(OBAdf, Excelrowcount):
    '''
    Update OBA.xlsx with data pulled from gasstorage.net
    '''
    # Count number of rows copied over
    obadatarows = len(OBAdf)

    # Change dataframe to chop off header and total row
    OBAdf2 = OBAdf[1:obadatarows-1]

    wb = openpyxl.load_workbook(loc2, keep_vba=True)
    ws = wb[sheetname]

    obalist = OBAdf2.values.tolist()
    for i in range(2, obadatarows):
        for j in range(1, 6):
                if j == 1:  # Catch Date column here
                    try:  # Convert to Datetime string
                        ws.cell(Excelrowcount+i, j).value = datetime.datetime.strptime(obalist[i-2][j-1],'%m/%d/%y')
                    except:  # Any errors converting datetime will just paste as string into Excel
                        ws.cell(Excelrowcount+i, j).value = obalist[i-2][j-1]
                else:
                    try:  # Convert string to float
                        a = obalist[i-2][j-1].replace(",", "")
                        ws.cell(Excelrowcount+i, j).value = int(a)
                    except:
                        ws.cell(Excelrowcount+i, j).value = obalist[i-2][j-1]

    wb.save(loc2)




if __name__ == '__main__':
    Excelrowcount, Begdate, Enddate = get_workbook_parameters(loc2, sheetname)
    OBAdf = df_from_web(Begdate, Enddate)
    update_workbook(OBAdf, Excelrowcount)
